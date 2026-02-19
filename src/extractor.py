import base64
import csv
import io
import logging
import os
from datetime import datetime

from src.utils import TEXT_EXTENSIONS, IMAGE_EXTENSIONS

logger = logging.getLogger("fanga")

MAX_TEXT_LENGTH = 1000


class FileExtractor:
    """Extract metadata and content from files."""

    def extract_metadata(self, filepath: str) -> dict:
        """Return file metadata dict."""
        stat = os.stat(filepath)
        size = stat.st_size
        return {
            "filename": os.path.basename(filepath),
            "extension": os.path.splitext(filepath)[1].lower(),
            "size_bytes": size,
            "size_human": self._human_size(size),
            "created_date": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d"),
            "modified_date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d"),
        }

    def extract_content(self, filepath: str) -> dict:
        """Extract readable content based on file type."""
        ext = os.path.splitext(filepath)[1].lower()
        try:
            if ext == ".pdf":
                return self._extract_pdf(filepath)
            elif ext == ".docx":
                return self._extract_docx(filepath)
            elif ext == ".xlsx":
                return self._extract_xlsx(filepath)
            elif ext == ".csv":
                return self._extract_csv(filepath)
            elif ext in IMAGE_EXTENSIONS:
                return self._extract_image(filepath)
            else:
                return {
                    "type": "text",
                    "content": os.path.basename(filepath),
                    "extraction_method": "filename_only",
                    "truncated": False,
                }
        except Exception as e:
            logger.error(f"Extraction failed for {filepath}: {e}")
            return {
                "type": "error",
                "content": "",
                "error": str(e),
            }

    def _extract_pdf(self, filepath: str) -> dict:
        import pdfplumber

        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[:2]:
                page_text = page.extract_text() or ""
                text += page_text + "\n"

        text = text.strip()

        # If text extraction yields < 20 chars, fall back to image
        if len(text) < 20:
            return self._extract_image(filepath)

        truncated = len(text) > MAX_TEXT_LENGTH
        if truncated:
            logger.warning(f"Content truncated for {filepath}")
            text = text[:MAX_TEXT_LENGTH]

        return {
            "type": "text",
            "content": text,
            "extraction_method": "pdfplumber",
            "truncated": truncated,
        }

    def _extract_docx(self, filepath: str) -> dict:
        from docx import Document

        doc = Document(filepath)
        text = "\n".join(p.text for p in doc.paragraphs)

        truncated = len(text) > MAX_TEXT_LENGTH
        if truncated:
            logger.warning(f"Content truncated for {filepath}")
            text = text[:MAX_TEXT_LENGTH]

        return {
            "type": "text",
            "content": text,
            "extraction_method": "python-docx",
            "truncated": truncated,
        }

    def _extract_xlsx(self, filepath: str) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True)
        lines = [f"Sheets: {', '.join(wb.sheetnames)}"]
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
        wb.close()

        text = "\n".join(lines)
        truncated = len(text) > MAX_TEXT_LENGTH
        if truncated:
            logger.warning(f"Content truncated for {filepath}")
            text = text[:MAX_TEXT_LENGTH]

        return {
            "type": "text",
            "content": text,
            "extraction_method": "openpyxl",
            "truncated": truncated,
        }

    def _extract_csv(self, filepath: str) -> dict:
        lines = []
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 11:
                    break
                lines.append(" | ".join(row))

        text = "\n".join(lines)
        truncated = len(text) > MAX_TEXT_LENGTH
        if truncated:
            logger.warning(f"Content truncated for {filepath}")
            text = text[:MAX_TEXT_LENGTH]

        return {
            "type": "text",
            "content": text,
            "extraction_method": "csv",
            "truncated": truncated,
        }

    def _extract_image(self, filepath: str) -> dict:
        with open(filepath, "rb") as f:
            data = base64.b64encode(f.read()).decode("utf-8")
        return {
            "type": "image",
            "content": data,
            "extraction_method": "vision",
            "truncated": False,
        }

    @staticmethod
    def _human_size(size_bytes: int) -> str:
        for unit in ("B", "KB", "MB", "GB"):
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"
